"""Excel解析模块 - 读取列名、Sheet信息、浮动图片定位"""

import os
import tempfile
import subprocess
from pathlib import Path
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class ImageInfo:
    """图片信息"""
    row: int          # Excel行号（1-based）
    col: int          # Excel列号（1-based）
    image_obj: object # openpyxl Image对象
    extension: str    # 图片扩展名（如 png, jpeg）


def convert_xls_to_xlsx(xls_path: str) -> str | None:
    """将.xls文件转换为.xlsx格式

    优先使用LibreOffice命令行转换，如果未安装则返回None

    Args:
        xls_path: .xls文件路径

    Returns:
        转换后的.xlsx文件路径，失败返回None
    """
    # 检查LibreOffice是否可用
    try:
        result = subprocess.run(
            ['soffice', '--version'],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            return None
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return None

    # 执行转换
    output_dir = tempfile.mkdtemp()
    try:
        result = subprocess.run(
            [
                'soffice',
                '--headless',
                '--convert-to', 'xlsx',
                '--outdir', output_dir,
                xls_path
            ],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            # 查找转换后的文件
            xlsx_name = Path(xls_path).stem + '.xlsx'
            xlsx_path = os.path.join(output_dir, xlsx_name)
            if os.path.exists(xlsx_path):
                return xlsx_path
    except (subprocess.TimeoutExpired, FileNotFoundError):
        pass

    return None


def get_sheet_names(file_path: str) -> list[str]:
    """获取Excel文件的所有Sheet名称

    Args:
        file_path: Excel文件路径

    Returns:
        Sheet名称列表
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    # 过滤掉WPS保留的Sheet
    sheets = [name for name in wb.sheetnames
              if not name.startswith('WpsReserved')]
    wb.close()
    return sheets


def get_headers(file_path: str, sheet_name: str) -> list[dict]:
    """获取指定Sheet的列名

    自动检测表头行：查找第一个有足够非空值的行

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称

    Returns:
        列信息列表 [{"index": 1, "name": "序号", "letter": "A"}, ...]
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # 在前10行中查找表头行（至少3个非空值的行）
    header_row = None
    for row in range(1, min(11, ws.max_row + 1)):
        non_empty = sum(
            1 for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value is not None
        )
        if non_empty >= 3:
            header_row = row
            break

    if header_row is None:
        header_row = 1

    # 读取表头
    headers = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None:
            name = str(value).replace('\n', ' ').strip()
            headers.append({
                "index": col,
                "name": name,
                "letter": get_column_letter(col)
            })

    wb.close()
    return headers


def get_header_row_index(file_path: str, sheet_name: str) -> int:
    """获取表头行的行号（1-based）"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    for row in range(1, min(11, ws.max_row + 1)):
        non_empty = sum(
            1 for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value is not None
        )
        if non_empty >= 3:
            wb.close()
            return row

    wb.close()
    return 1


def locate_images(file_path: str, sheet_name: str) -> list[ImageInfo]:
    """定位Excel中的所有浮动图片

    通过图片anchor坐标确定图片所在的行列

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称

    Returns:
        ImageInfo列表
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    images = []
    for img in ws._images:
        anchor = img.anchor

        # 获取起始行列（anchor是0-based，转为1-based）
        if hasattr(anchor, '_from'):
            row = anchor._from.row + 1
            col = anchor._from.col + 1
        else:
            continue

        # 获取图片扩展名
        ext = _get_image_extension(img)

        images.append(ImageInfo(
            row=row,
            col=col,
            image_obj=img,
            extension=ext
        ))

    wb.close()
    return images


def get_image_data(image_info: ImageInfo) -> bytes:
    """获取图片的二进制数据

    Args:
        image_info: 图片信息对象

    Returns:
        图片二进制数据
    """
    return image_info.image_obj._data()


def get_cell_value(file_path: str, sheet_name: str, row: int, col: int) -> str | None:
    """获取指定单元格的值

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称
        row: 行号（1-based）
        col: 列号（1-based）

    Returns:
        单元格值（字符串），空则为None
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    value = ws.cell(row=row, column=col).value
    wb.close()

    if value is None:
        return None
    return str(value).strip()


def _get_image_extension(img) -> str:
    """获取图片的扩展名（不读取图片数据，避免服务器上句柄问题）"""
    # 尝试从图片格式获取
    if hasattr(img, 'format') and img.format:
        return img.format.lower()

    # 从 img.path 的后缀猜测
    path = getattr(img, 'path', '')
    if path:
        ext = os.path.splitext(path)[1].lower().lstrip('.')
        if ext in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp', 'ico', 'emf', 'wmf'):
            return ext if ext != 'jpeg' else 'jpg'

    # 默认png
    return 'png'
