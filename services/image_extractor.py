"""图片提取模块 - 根据配置提取图片、应用命名规则、打包ZIP"""

import os
import re
import uuid
import zipfile
import shutil
import tempfile
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .file_sanitizer import build_filename, resolve_duplicates


@dataclass
class ExtractConfig:
    """提取配置"""
    file_path: str          # Excel文件路径
    sheet_name: str         # Sheet名称
    image_col: int          # 图片所在列（1-based）
    name_col: int           # 商品名称列（1-based）
    prefix_col: int | None  # 前缀字段列（可选，1-based）
    manual_prefix: str      # 手动前缀（可选）


@dataclass
class ExtractResult:
    """提取结果"""
    task_id: str            # 任务ID
    total_images: int       # 总图片数
    extracted: int          # 成功提取数
    skipped: int            # 跳过数（无名称等）
    zip_path: str | None    # ZIP文件路径
    errors: list[str]       # 错误信息列表


def _detect_header_row(ws) -> int:
    """检测表头行号（1-based），查找前10行中非空值>=3的行"""
    for row in range(1, min(11, ws.max_row + 1)):
        non_empty = sum(
            1 for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value is not None
        )
        if non_empty >= 3:
            return row
    return 1


def _get_image_extension(data: bytes) -> str:
    """通过文件头判断图片格式"""
    if data[:8] == b'\x89PNG\r\n\x1a\n':
        return 'png'
    elif data[:2] == b'\xff\xd8':
        return 'jpg'
    elif data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return 'webp'
    elif data[:6] in (b'GIF87a', b'GIF89a'):
        return 'gif'
    elif data[:4] == b'\x00\x00\x01\x00':
        return 'ico'
    elif data[:2] == b'BM':
        return 'bmp'
    elif data[:4] == b'\x01\x00\x00\x00':
        return 'emf'
    elif data[:4] == b'\xd7\xcd\xc6\x9a':
        return 'wmf'
    return 'png'


def _get_cell_value(ws, row: int, col: int) -> str | None:
    """获取单元格值，自动处理合并单元格（取合并区域主单元格值）"""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return str(cell.value).strip()

    # 检查该单元格是否位于某个合并区域中
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            main_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if main_val is not None:
                return str(main_val).strip()
    return None


def extract_images(config: ExtractConfig, output_dir: str | None = None) -> ExtractResult:
    """执行图片提取（只加载一次Excel文件）

    Args:
        config: 提取配置
        output_dir: 输出目录，None则使用临时目录

    Returns:
        提取结果
    """
    task_id = str(uuid.uuid4())[:8]
    errors = []

    if output_dir is None:
        output_dir = tempfile.mkdtemp(prefix=f"excel_img_{task_id}_")

    extract_dir = os.path.join(output_dir, "images")
    os.makedirs(extract_dir, exist_ok=True)

    # 只加载一次Excel文件
    wb = load_workbook(config.file_path)
    ws = wb[config.sheet_name]

    # 同时打开xlsx作为zip包，直接读取图片原始字节（避免img._data()依赖内部句柄）
    xlsx_zip = zipfile.ZipFile(config.file_path, 'r')

    header_row = _detect_header_row(ws)

    # 构建媒体文件名映射（img.path扩展名不可靠，用实际zip内文件名匹配）
    media_map = {}
    for name in xlsx_zip.namelist():
        if name.startswith('xl/media/'):
            base = os.path.splitext(os.path.basename(name))[0]  # image1
            media_map[base] = name

    # 定位图片：获取anchor坐标及覆盖范围（处理悬浮跨单元格的情况）
    target_images = []
    path_debug = []
    for img in ws._images:
        anchor = img.anchor
        if not hasattr(anchor, '_from'):
            continue

        from_row = anchor._from.row + 1  # 0-based → 1-based
        from_col = anchor._from.col + 1

        # 获取结束行列（TwoCellAnchor有_to属性，表示图片覆盖到的单元格）
        to_row = from_row
        to_col = from_col
        if hasattr(anchor, '_to'):
            to_row = anchor._to.row + 1
            to_col = anchor._to.col + 1

        # 判断图片是否关联到目标图片列：
        # 1. 图片覆盖范围包含目标列（处理悬浮超长的情况）
        # 2. 或anchor起始列在目标列附近（±1容错）
        col_min = min(from_col, to_col)
        col_max = max(from_col, to_col)
        col_in_range = col_min <= config.image_col <= col_max
        col_nearby = abs(from_col - config.image_col) <= 1

        # 只保留数据行（在表头之后）
        if not (col_in_range or col_nearby) or from_row <= header_row:
            continue

        # 优先使用 img._data() 直接读取图片原始字节（绕过 img.path 解析错误）
        img_data = None
        ext = 'png'
        try:
            img_data = img._data()
            ext = _get_image_extension(img_data)
        except Exception:
            # 备用：通过 img.path 从 zip 包读取（记录 path 用于调试）
            try:
                base_name = os.path.splitext(os.path.basename(img.path))[0]
                path_debug.append(base_name)
                zip_path = media_map.get(base_name)
                if zip_path:
                    img_data = xlsx_zip.read(zip_path)
                    ext = _get_image_extension(img_data)
            except Exception:
                pass

        target_images.append((from_row, from_col, ext, img, img_data, to_row, to_col))

    # 调试：检查 img.path 是否全部相同（如果用了备用路径）
    if path_debug and len(set(path_debug)) == 1 and len(path_debug) > 1:
        errors.append(
            f"[调试警告] 检测到 {len(path_debug)} 张图片的 img.path 全部相同（{path_debug[0]}），"
            f"已尝试用 img._data() 绕过，如果图片仍相同请反馈"
        )

    if not target_images:
        xlsx_zip.close()
        wb.close()
        return ExtractResult(
            task_id=task_id,
            total_images=len(ws._images),
            extracted=0,
            skipped=len(ws._images),
            zip_path=None,
            errors=["在指定图片列中未找到图片，请确认图片列是否正确"]
        )

    # 按Excel行号排序，确保导出顺序与表格一致
    target_images.sort(key=lambda t: (t[0], t[1]))

    # 根据最大行号动态计算零填充宽度，保证超过999行也能正确排序
    max_row = max(t[0] for t in target_images) if target_images else 0
    row_pad_width = len(str(max_row))

    # 预读所有需要的单元格值（支持合并单元格）
    rows_needed = list(set(t[0] for t in target_images))
    cols_needed = [config.name_col]
    if config.prefix_col:
        cols_needed.append(config.prefix_col)

    cell_cache = {}
    for row in rows_needed:
        for col in cols_needed:
            cell_cache[(row, col)] = _get_cell_value(ws, row, col)

    # 构建文件名列表（同时保留图片数据和原始名称）
    image_data_pairs = []
    raw_names = []
    debug_info = []
    for idx, (row, col, ext, img, img_data, to_row, to_col) in enumerate(target_images):
        name_value = cell_cache.get((row, config.name_col))

        # 容错：如果直接读取为空，尝试在图片覆盖的行范围内查找名称
        if not name_value and to_row > row:
            for r in range(row, to_row + 1):
                name_value = _get_cell_value(ws, r, config.name_col)
                if name_value:
                    break

        field_prefix = ""
        if config.prefix_col:
            field_prefix = cell_cache.get((row, config.prefix_col)) or ""

        # 记录调试信息：哪些图片读取到了空名称
        if not name_value:
            debug_info.append(
                f"第{row}行图片名称列为空，anchor覆盖范围="
                f"({row},{col})->({to_row},{to_col})"
            )

        filename = build_filename(
            name_col_value=name_value or "",
            ext=ext,
            manual_prefix=config.manual_prefix,
            field_prefix_value=field_prefix,
            fallback_index=idx + 1,
            row_num=row,
            row_pad_width=row_pad_width,
        )
        image_data_pairs.append((filename, img_data))
        raw_names.append(name_value or "")

    # 输出调试信息（最多5条，避免刷屏）
    if debug_info:
        errors.append(f"[调试] 发现 {len(debug_info)} 个图片未读取到名称，前5条：")
        errors.extend(debug_info[:5])

    # 解决重复文件名
    filenames = resolve_duplicates([f[0] for f in image_data_pairs])

    # 写入图片文件（优先使用 img._data() 获取的字节，避免 img.path 解析错误）
    extracted = 0
    skipped = 0
    name_mapping = []

    for i, ((filename, img_data), final_filename) in enumerate(zip(image_data_pairs, filenames)):
        filepath = os.path.join(extract_dir, final_filename)
        original_name = raw_names[i]
        if not img_data:
            skipped += 1
            errors.append(f"提取失败: {final_filename} - 无法读取图片数据（img._data() 和 zip 备用均失败）")
            continue
        try:
            with open(filepath, 'wb') as f:
                f.write(img_data)
            extracted += 1
            name_mapping.append((original_name, final_filename))
        except Exception as e:
            skipped += 1
            errors.append(f"提取失败: {final_filename} - {str(e)}")

    # 生成文件名映射CSV（RPA直接用这个做匹配）
    if name_mapping:
        import csv
        mapping_path = os.path.join(extract_dir, "mapping.csv")
        with open(mapping_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['original_name', 'exported_filename'])
            for original, exported in name_mapping:
                writer.writerow([original, exported])
        errors.insert(0, f"[成功] 已生成 mapping.csv（共 {len(name_mapping)} 条记录），已包含在ZIP包中")
        print(f"[DEBUG] mapping.csv generated: {mapping_path}, rows={len(name_mapping)}")

    # 关闭资源
    xlsx_zip.close()
    wb.close()

    # 打包ZIP
    zip_path = None
    if extracted > 0:
        zip_filename = f"images_{task_id}.zip"
        zip_path = os.path.join(output_dir, zip_filename)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname in os.listdir(extract_dir):
                fpath = os.path.join(extract_dir, fname)
                zf.write(fpath, fname)
        shutil.rmtree(extract_dir, ignore_errors=True)

    return ExtractResult(
        task_id=task_id,
        total_images=len(target_images),
        extracted=extracted,
        skipped=skipped,
        zip_path=zip_path,
        errors=errors
    )
